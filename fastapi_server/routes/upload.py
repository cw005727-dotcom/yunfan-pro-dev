"""文件上传路由"""
import os
import uuid
import asyncio
import tempfile
import time
import json
import base64
import sqlite3
import openpyxl
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from ..config import UPLOAD_DIR, DB_PATH, DATA_DIR
from ..middleware.auth import get_ml_token_provider

# 导入解析脚本
import sys
_scripts_path = os.path.join(os.path.dirname(__file__), '..', '..', 'scripts')
sys.path.insert(0, _scripts_path)

router = APIRouter(prefix="/api", tags=["上传"])


@router.post("/upload/orders")
async def upload_orders(file: UploadFile = File(...)):
    """上传订单Excel，解析入库到 operational_orders 表"""
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="文件过大（最大20MB）")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        from upload.parse_orders import parse_and_import as parse_orders_excel
        result = parse_orders_excel(tmp_path, source_file=file.filename)
        if result.get('error'):
            raise HTTPException(status_code=422, detail=result['error'])
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")
    finally:
        os.unlink(tmp_path)


@router.post("/upload/links")
async def upload_links(
    file: UploadFile = File(...),
    site_id: str = "MLB",
):
    """上传商品性能Excel，解析入库到 product_performance 表
    解析后立即返回，图片在后台上异步拉取，不阻塞请求。
    """
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="文件过大（最大20MB）")

    site_id = site_id.upper()

    # 解析 XLSX
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active if '报告' not in wb.sheetnames else wb['报告']

        headers = []
        for row in ws.iter_rows(min_row=6, max_row=6, values_only=True):
            headers = list(row)

        field_map = {
            '发布 ID': 'item_id', '商品': 'product_name', '当前状态': 'status',
            '变体': 'variation', 'SKU': 'sku', '独立访问量': 'unique_visits',
            '订单数量': 'order_count', '唯一买家': 'unique_buyers',
            '已售件数': 'units_sold', '毛销售额 (USD)': 'gross_sales_usd',
            '占比(%)': 'share_percent', '访客转化率': 'visitor_convert_rate',
            '访客到购买的转化率': 'visitor_buy_convert_rate'
        }

        conn = sqlite3.connect(str(DB_PATH))

        # 第一阶段：只解析数据，快速入库（不带图片）
        item_ids = []
        imported = 0
        skipped = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
            item_id_raw = row[0]
            if not item_id_raw or not str(item_id_raw).isdigit():
                continue
            item_id = str(item_id_raw)
            item_ids.append(item_id)

            data = {'item_id': item_id, 'site_id': site_id}
            for col_idx, value in enumerate(row):
                header = headers[col_idx] if col_idx < len(headers) else None
                if header and header in field_map:
                    key = field_map[header]
                    data[key] = value

            for num_field in ['unique_visits', 'order_count', 'unique_buyers', 'units_sold', 'gross_sales_usd']:
                if num_field in data and data[num_field]:
                    val = str(data[num_field]).replace(',', '.').replace('US$', '').strip()
                    try:
                        data[num_field] = float(val) if '.' in val else int(val)
                    except:
                        data[num_field] = 0

            data['source_file'] = file.filename

            try:
                conn.execute("""
                    INSERT INTO product_performance 
                    (item_id, sku, product_name, status, variation, unique_visits, order_count,
                     unique_buyers, units_sold, gross_sales_usd, share_percent, 
                     visitor_convert_rate, visitor_buy_convert_rate, source_file, site_id, updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)
                    ON CONFLICT(item_id) DO UPDATE SET
                        sku=excluded.sku, product_name=excluded.product_name, status=excluded.status,
                        variation=excluded.variation, unique_visits=excluded.unique_visits,
                        order_count=excluded.order_count, unique_buyers=excluded.unique_buyers,
                        units_sold=excluded.units_sold, gross_sales_usd=excluded.gross_sales_usd,
                        share_percent=excluded.share_percent, visitor_convert_rate=excluded.visitor_convert_rate,
                        visitor_buy_convert_rate=excluded.visitor_buy_convert_rate,
                        source_file=excluded.source_file, site_id=excluded.site_id,
                        updated_at=CURRENT_TIMESTAMP
                """, (
                    data.get('item_id'), data.get('sku'), data.get('product_name'),
                    data.get('status', ''), data.get('variation', ''),
                    data.get('unique_visits', 0), data.get('order_count', 0),
                    data.get('unique_buyers', 0), data.get('units_sold', 0),
                    data.get('gross_sales_usd', 0), data.get('share_percent', ''),
                    data.get('visitor_convert_rate', ''), data.get('visitor_buy_convert_rate', ''),
                    file.filename, site_id
                ))
                imported += 1
            except Exception:
                skipped += 1

        conn.commit()
        conn.close()

        # 第二阶段：后台拉取图片（不阻塞用户返回）
        if imported > 0 and item_ids:
            import threading
            _provider = get_ml_token_provider()
            def _pull_images(item_ids, site_id, tmp_path_for_cleanup):
                import requests, json, time, sqlite3
                conn = sqlite3.connect(str(DB_PATH))
                for idx, item_id in enumerate(item_ids):
                    try:
                        _provider.clear_cache()
                        token = _provider.get_valid_token()
                        if not token:
                            continue
                        ml_id = f'{site_id}{item_id}'
                        r = requests.get(
                            f'https://api.mercadolibre.com/marketplace/items/{ml_id}',
                            headers={'Authorization': f'Bearer {token}'}, timeout=8
                        )
                        if r.status_code == 200:
                            d = r.json()
                            thumbnail = d.get('thumbnail', '') or d.get('secure_thumbnail', '')
                            pics = d.get('pictures', [])
                            pics_count = len(pics)
                            pics_urls = [p if isinstance(p, str) else p.get('url', '') for p in pics]
                            pictures = json.dumps(pics_urls, ensure_ascii=False)
                            conn.execute(
                                "UPDATE product_performance SET thumbnail=?, pictures=?, pictures_count=? WHERE item_id=?",
                                (thumbnail, pictures, pics_count, item_id)
                            )
                            conn.commit()
                    except:
                        pass
                    time.sleep(0.3)
                conn.close()
                try:
                    os.unlink(tmp_path)
                except:
                    pass

            t = threading.Thread(target=_pull_images, args=(item_ids, site_id, tmp_path), daemon=True)
            t.start()
        else:
            os.unlink(tmp_path)

        return {
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'message': f'导入成功 {imported} 条，跳过 {skipped} 条'
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")
    finally:
        # 如果后台线程已经接手清理，这里不重复删除
        try:
            os.unlink(tmp_path)
        except:
            pass


@router.post("/upload/logistics")
async def upload_logistics(file: UploadFile = File(...)):
    """上传物流追踪Excel"""
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="文件过大（最大20MB）")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=422, detail="Excel为空")

        headers = list(rows[0])
        field_map = {
            '订单号': 'order_number',
            '站点': 'site',
            '自定义店铺名': 'store_name',
            '下单时间': 'order_date',
            '订单状态': 'status',
            'listing ID': 'listing_id',
            'SKU': 'sku',
            '1688订单号': 'logistics_1688_order',
            '物流单号': 'logistics_1688_tracking',
            '贴单状态': 'label_status',
            '入仓时间': 'warehouse_in_date',
            '国际物流单号': 'international_tracking',
        }

        col_idx = {}
        for i, h in enumerate(headers):
            if h and h in field_map:
                col_idx[field_map[h]] = i

        if 'order_number' not in col_idx:
            raise HTTPException(status_code=422, detail="Excel缺少「订单号」列")

        conn = sqlite3.connect(str(DB_PATH))
        imported = 0
        skipped = 0

        for row in rows[1:]:
            try:
                order_number = str(row[col_idx['order_number']]).strip() if row[col_idx['order_number']] else ''
                if not order_number or not order_number.isdigit():
                    skipped += 1
                    continue

                data = {'order_number': order_number, 'is_ignored': 0}
                for field, idx in col_idx.items():
                    if field == 'order_number':
                        continue
                    val = row[idx] if idx < len(row) else None
                    if val is not None:
                        data[field] = str(val).strip()

                conn.execute("""
                    INSERT INTO logistics_tracking
                    (order_number, site, store_name, order_date, status,
                     listing_id, sku,
                     logistics_1688_order, logistics_1688_tracking, label_status,
                     warehouse_in_date, international_tracking, is_ignored)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,0)
                    ON CONFLICT(order_number) DO UPDATE SET
                        site=excluded.site, store_name=excluded.store_name,
                        order_date=excluded.order_date, status=excluded.status,
                        listing_id=excluded.listing_id, sku=excluded.sku,
                        logistics_1688_order=excluded.logistics_1688_order,
                        logistics_1688_tracking=excluded.logistics_1688_tracking,
                        label_status=excluded.label_status,
                        warehouse_in_date=excluded.warehouse_in_date,
                        international_tracking=excluded.international_tracking,
                        is_ignored=excluded.is_ignored
                """, (
                    data.get('order_number'), data.get('site'), data.get('store_name'),
                    data.get('order_date'), data.get('status'),
                    data.get('listing_id'), data.get('sku'),
                    data.get('logistics_1688_order'),
                    data.get('logistics_1688_tracking'), data.get('label_status'),
                    data.get('warehouse_in_date'), data.get('international_tracking'),
                ))
                imported += 1
            except Exception:
                skipped += 1

        conn.commit()
        conn.close()
        os.unlink(tmp_path)

        return {'success': True, 'imported': imported, 'skipped': skipped,
                'message': f'导入成功 {imported} 条，跳过 {skipped} 条'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")
    finally:
        try:
            os.unlink(tmp_path)
        except:
            pass


@router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """通用文件上传"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename")
    ext = os.path.splitext(file.filename)[1]
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = UPLOAD_DIR / filename
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 20MB)")
    await asyncio.to_thread(lambda f=filepath, d=content: open(f, 'wb').write(d))
    return {"url": f"/uploads/{filename}", "filename": filename}


@router.get("/uploads/{filename}")
async def serve_upload(filename: str):
    filepath = UPLOAD_DIR / filename
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(filepath)
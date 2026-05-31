#!/usr/bin/env python3
"""
解析运营系统导出的订单Excel，upsert到operational_orders表
唯一键：order_number（编号）
"""

import sys
import openpyxl
import sqlite3
import re
from datetime import datetime
from pathlib import Path

# Excel列名 → DB字段名 映射
COLUMN_MAP = {
    "订单号": "order_number",
    "编号": "order_number",
    "时间": "order_date",
    "业务员": "salesperson",
    "来源": "source",
    "公司店铺": "store",
    "订单状态": "status",
    "状态": "status",
    "金额": "amount_usd",
    "费用": "fee",
    "税费": "tax",
    "退款": "refund",
    "人民币收入": "amount_cny",
    "采购成本": "purchase_cost",
    "利润": "profit",
    "结余": "balance",
    "产品id": "product_id",
    "ASIN": "asin",
    "SKU": "sku",
    "标题": "product_name",
    "中文简称": "product_short_name",
    "图片": "thumbnail",
    "毛重": "weight",
    "尺寸": "dimensions",
    "数量": "quantity",
    "物流": "logistics",
    "渠道": "channel",
    "运单号": "waybill_no",
    "追踪号": "tracking_no",
    "运输商": "carrier",
    "运费": "shipping_fee",
    "海外仓退货费": "overseas_warehouse_return_fee",
    "订单运费": "order_shipping_fee",
    "订单备注": "order_remark",
    "定制单备注": "custom_order_remark",
    "地区": "site",
    "买家名称": "buyer_name",
    "邮箱": "email",
    "电话": "phone",
    "城市": "city",
    "州省": "state",
    "邮编": "zipcode",
    "地址": "address",
    "备货时间": "prepare_time",
    "送达时间": "delivery_time",
    "物流费": "logistics_fee",
    "币种": "currency",
    "采购单号": "purchase_order_no",
    "采购追踪": "purchase_tracking",
    "采购链接": "purchase_link",
    "产品供应货号": "supply_part_no",
}

DB_PATH = Path(__file__).parent.parent.parent / "mercadolibre.db"


from typing import Optional, Union

def col_to_db_key(col_name: str) -> Optional[str]:
    """标准化列名，转为DB字段名"""
    col = col_name.strip()
    return COLUMN_MAP.get(col)


def parse_value(val, db_key: Union[str, tuple]):
    """根据字段类型做清洗"""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ("None", "nan", ""):
        return None

    # 数值字段
    if db_key in (
        "amount_usd", "fee", "tax", "refund", "amount_cny",
        "purchase_cost", "profit", "balance", "weight",
        "shipping_fee", "logistics_fee", "overseas_warehouse_return_fee",
        "order_shipping_fee", "quantity"
    ):
        # 去掉 $ , 等字符
        s = re.sub(r"[\$,]", "", s)
        try:
            return float(s)
        except:
            return None

    # 整数字段
    if db_key in ("quantity",):
        try:
            return int(float(s))
        except:
            return None

    return s


def get_header_map(headers: tuple) -> dict[str, str]:
    """返回 {excel_col_index: db_key}"""
    result = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        db_key = col_to_db_key(str(h))
        if db_key:
            result[i] = db_key
    return result


def upsert_order(cursor, row_data: dict, source_file: str, conn=None):
    """Upsert一条订单，unique key = order_number，同时记录 profit/status 变化"""
    order_number = row_data.get("order_number")
    row_data["source_file"] = source_file
    row_data["updated_at"] = datetime.now().isoformat()

    # 对比旧数据，记录 profit/status 变化
    if order_number and conn:
        old_row = cursor.execute(
            "SELECT profit, status, COALESCE(thumbnail,''), COALESCE(site,''), COALESCE(store,'') FROM operational_orders WHERE order_number = ?",
            (order_number,)
        ).fetchone()
        if old_row:
            old_profit, old_status = old_row[0], old_row[1]
            new_profit = row_data.get("profit")
            new_status = row_data.get("status")
            thumbnail = row_data.get("thumbnail") or old_row[2]
            site = row_data.get("site") or old_row[3]
            store_name = row_data.get("store") or old_row[4]

            # profit 变化 → 利润变化卡片
            if str(old_profit or "") != str(new_profit or ""):
                cursor.execute("""
                    INSERT INTO order_changes (order_number, change_type, old_value, new_value, thumbnail, site, store_name)
                    VALUES (?, 'profit', ?, ?, ?, ?, ?)
                """, (order_number, str(old_profit or ""), str(new_profit or ""), thumbnail, site, store_name))
            # status 变化（已采购→在途中等） → 物流变化卡片
            if str(old_status or "") != str(new_status or ""):
                cursor.execute("""
                    INSERT INTO order_changes (order_number, change_type, old_value, new_value, thumbnail, site, store_name)
                    VALUES (?, 'logistics', ?, ?, ?, ?, ?)
                """, (order_number, str(old_status or ""), str(new_status or ""), thumbnail, site, store_name))

    placeholders = ", ".join(["?"] * len(row_data))
    col_names = ", ".join(row_data.keys())

    # INSERT OR REPLACE：已存在则替换
    sql = f"""
        INSERT OR REPLACE INTO operational_orders ({col_names})
        VALUES ({placeholders})
    """
    cursor.execute(sql, list(row_data.values()))


def parse_and_import(xlsx_path: str, source_file: str = None, owner: str = None) -> dict:
    """解析Excel并导入数据库，返回统计"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    if ws.title != "订单":
        # 找包含"订单"的sheet
        for s in wb.sheetnames:
            if "订单" in s:
                ws = wb[s]
                break

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"error": "Excel为空"}

    headers = rows[0]
    header_map = get_header_map(headers)

    if not header_map:
        return {"error": f"未识别到有效列，请检查表头。首行: {headers[:10]}"}

    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            row_data = {"owner_username": owner or ""}
            for col_idx, db_key in header_map.items():
                if col_idx >= len(row):
                    continue
                val = parse_value(row[col_idx], db_key)
                if val is not None:
                    row_data[db_key] = val

            if not row_data.get("order_number"):
                skipped += 1
                continue

            upsert_order(cursor, row_data, source_file or xlsx_path, conn)
            imported += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {e}")

    conn.commit()
    conn.close()

    return {
        "total": len(rows) - 1,
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:20],  # 最多返回20条
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python parse_orders.py <订单.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    source_file = Path(xlsx_path).name

    print(f"解析文件: {source_file}")
    result = parse_and_import(xlsx_path, source_file)

    print(f"\n导入结果:")
    print(f"  总行数:   {result.get('total', 0)}")
    print(f"  成功:     {result.get('imported', 0)}")
    print(f"  跳过:     {result.get('skipped', 0)}")
    if result.get("errors"):
        print(f"  错误:")
        for e in result["errors"]:
            print(f"    {e}")
    else:
        print(f"  错误:     无")
